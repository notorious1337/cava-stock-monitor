import os
import json
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from urllib.parse import urljoin
from datetime import datetime

import requests
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from email.mime.base import MIMEBase
from email import encoders

# ====== CONFIG (from environment) ======

BASE_URL = "https://cavaathleisure.com"

SMTP_SERVER = os.getenv("SMTP_SERVER", "smtp.gmail.com")
SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
SMTP_USER = os.getenv("SMTP_USER")          # FROM email
SMTP_PASSWORD = os.getenv("SMTP_PASSWORD")  # SMTP / app password
TO_EMAIL = os.getenv("TO_EMAIL", SMTP_USER) # Default TO == FROM if not set

# Email only if there is ANY change compared to previous day
ONLY_EMAIL_IF_CHANGES = os.getenv("ONLY_EMAIL_IF_CHANGES", "true").lower() == "true"

STATE_FILE = "state.json"  # stored in repo to compare day-to-day


# ====== HTTP HELPERS ======

def fetch_products_page(page: int, limit: int = 250):
    """
    Fetch a single page of products from Shopify /products.json.
    Returns list of products or [] if none.
    """
    url = f"{BASE_URL}/products.json?limit={limit}&page={page}"
    headers = {
        "User-Agent": "Mozilla/5.0 (compatible; CavaStockBot/1.0)",
        "Accept": "application/json",
    }
    resp = requests.get(url, headers=headers, timeout=30)
    if resp.status_code != 200:
        print(f"Got status {resp.status_code} for {url}, stopping pagination.")
        return []
    data = resp.json()
    return data.get("products", [])


def fetch_all_products():
    """
    Fetch all products across all /products.json pages.
    """
    all_products = []
    page = 1

    while True:
        products = fetch_products_page(page)
        if not products:
            break
        all_products.extend(products)
        print(f"Fetched {len(products)} products from page {page}")
        page += 1

    print(f"Total products fetched: {len(all_products)}")
    return all_products


# ====== BUILD REPORT FROM PRODUCTS.JSON ======

def build_report_via_products_json():
    """
    Build a report mapping, with three buckets:

    {
      "partial": {
        product_url: {
          "title": ...,
          "available_sizes": [...],
          "unavailable_sizes": [...]
        },
        ...
      },
      "full_in_stock": { ... },
      "full_oos": { ... }
    }
    """
    partial = {}
    full_in_stock = {}
    full_oos = {}

    products = fetch_all_products()

    for p in products:
        title = p.get("title", "Unknown product")
        handle = p.get("handle")
        if not handle:
            continue

        product_url = urljoin(BASE_URL, f"/products/{handle}")

        available_sizes = []
        unavailable_sizes = []
        total_variants_with_size = 0

        for v in p.get("variants", []):
            available = v.get("available", True)
            size = v.get("option1") or v.get("title")
            if not size:
                continue
            total_variants_with_size += 1
            if available:
                available_sizes.append(size)
            else:
                unavailable_sizes.append(size)

        if total_variants_with_size == 0:
            # No meaningful size info; skip
            continue

        unique_available = sorted(set(available_sizes))
        unique_unavailable = sorted(set(unavailable_sizes))

        # Classify product
        if unique_available and unique_unavailable:
            # Partially sold out
            partial[product_url] = {
                "title": title,
                "available_sizes": unique_available,
                "unavailable_sizes": unique_unavailable,
            }
            print(f"[PARTIAL OOS] {title} -> OOS: {', '.join(unique_unavailable)} | IN STOCK: {', '.join(unique_available)}")
        elif unique_available and not unique_unavailable:
            # Fully available
            full_in_stock[product_url] = {
                "title": title,
                "available_sizes": unique_available,
                "unavailable_sizes": [],
            }
            print(f"[FULL IN STOCK] {title} -> {', '.join(unique_available)}")
        elif unique_unavailable and not unique_available:
            # Fully sold out
            full_oos[product_url] = {
                "title": title,
                "available_sizes": [],
                "unavailable_sizes": unique_unavailable,
            }
            print(f"[FULL OOS] {title} -> ALL SIZES OOS: {', '.join(unique_unavailable)}")
        else:
            # Weird edge case
            continue

    return {
        "partial": partial,
        "full_in_stock": full_in_stock,
        "full_oos": full_oos,
    }


# ====== HTML & TEXT REPORT FORMATTERS ======

def render_table_html(title, items, status):
    """
    status = 'partial' | 'full_in_stock' | 'full_oos'
    """
    if status == "partial":
        header_color = "#ffb347"  # orange-ish
    elif status == "full_in_stock":
        header_color = "#77dd77"  # green-ish
    else:
        header_color = "#ff6961"  # red-ish

    if not items:
        return f"""
        <h2>{title}</h2>
        <p style="color:#666;">None.</p>
        """

    rows_html = ""
    for url, data in items.items():
        name = data["title"]
        available = ", ".join(data["available_sizes"]) if data["available_sizes"] else "None"
        unavailable = ", ".join(data["unavailable_sizes"]) if data["unavailable_sizes"] else "None"

        if status == "partial":
            status_label = "Partially Sold-Out"
        elif status == "full_in_stock":
            status_label = "Fully Available"
        else:
            status_label = "Fully Sold-Out"

        rows_html += f"""
        <tr>
            <td style="padding:8px; border:1px solid #ddd; font-weight:500;">
                <a href="{url}" style="color:#1a73e8; text-decoration:none;">{name}</a>
            </td>
            <td style="padding:8px; border:1px solid #ddd;">{status_label}</td>
            <td style="padding:8px; border:1px solid #ddd;">{available}</td>
            <td style="padding:8px; border:1px solid #ddd;">{unavailable}</td>
        </tr>
        """

    return f"""
    <h2 style="margin-top:24px; margin-bottom:8px;">{title}</h2>
    <table cellpadding="0" cellspacing="0" style="border-collapse:collapse; width:100%; max-width:900px; font-size:14px;">
        <thead>
            <tr style="background:{header_color}; color:#000;">
                <th style="padding:8px; border:1px solid #ddd; text-align:left;">Product</th>
                <th style="padding:8px; border:1px solid #ddd; text-align:left;">Status</th>
                <th style="padding:8px; border:1px solid #ddd; text-align:left;">Available Sizes</th>
                <th style="padding:8px; border:1px solid #ddd; text-align:left;">Sold-Out Sizes</th>
            </tr>
        </thead>
        <tbody>
            {rows_html}
        </tbody>
    </table>
    """


def format_report_html(report):
    partial = report.get("partial", {})
    full_in_stock = report.get("full_in_stock", {})
    full_oos = report.get("full_oos", {})

    partial_html = render_table_html("1) Partially Sold-Out Products", partial, "partial")
    full_in_stock_html = render_table_html("2) Fully Available Products", full_in_stock, "full_in_stock")
    full_oos_html = render_table_html("3) Fully Sold-Out Products", full_oos, "full_oos")

    html = f"""
    <html>
    <head>
      <meta charset="UTF-8" />
      <title>CAVA Daily Stock Report</title>
    </head>
    <body style="font-family:Arial,Helvetica,sans-serif; background:#f5f5f5; margin:0; padding:16px;">
      <div style="max-width:960px; margin:0 auto; background:#ffffff; padding:16px 24px; border-radius:8px; box-shadow:0 2px 4px rgba(0,0,0,0.08);">
        <h1 style="font-size:20px; margin-bottom:4px;">CAVA Daily Stock Report</h1>
        <p style="margin-top:0; color:#666; font-size:13px;">Source: /products.json • Generated automatically</p>
        {partial_html}
        {full_in_stock_html}
        {full_oos_html}
        <p style="margin-top:24px; font-size:11px; color:#999;">
          This email was generated by an automated GitHub Actions job.
        </p>
      </div>
    </body>
    </html>
    """
    return html


def format_report_text(report):
    """
    Simple plain-text fallback for email clients that don't support HTML.
    """
    partial = report.get("partial", {})
    full_in_stock = report.get("full_in_stock", {})
    full_oos = report.get("full_oos", {})

    lines = []
    lines.append("CAVA Daily Stock Report")
    lines.append("Source: /products.json")
    lines.append("")

    lines.append("1) Partially Sold-Out Products")
    lines.append("--------------------------------")
    if not partial:
        lines.append("None.")
    else:
        for url, data in partial.items():
            lines.append(f"[PARTIAL] {data['title']}")
            lines.append(f"  URL: {url}")
            lines.append(f"  Available sizes: {', '.join(data['available_sizes']) or 'None'}")
            lines.append(f"  Sold-out sizes: {', '.join(data['unavailable_sizes']) or 'None'}")
            lines.append("")
    lines.append("")

    lines.append("2) Fully Available Products")
    lines.append("--------------------------------")
    if not full_in_stock:
        lines.append("None.")
    else:
        for url, data in full_in_stock.items():
            lines.append(f"[FULL IN STOCK] {data['title']}")
            lines.append(f"  URL: {url}")
            lines.append(f"  All sizes: {', '.join(data['available_sizes']) or 'None'}")
            lines.append("")
    lines.append("")

    lines.append("3) Fully Sold-Out Products")
    lines.append("--------------------------------")
    if not full_oos:
        lines.append("None.")
    else:
        for url, data in full_oos.items():
            lines.append(f"[FULL OOS] {data['title']}")
            lines.append(f"  URL: {url}")
            lines.append(f"  All sold-out sizes: {', '.join(data['unavailable_sizes']) or 'None'}")
            lines.append("")

    return "\n".join(lines)


# ====== STATE / DIFF LOGIC ======

def load_previous_state():
    """
    Load previous snapshot from state.json.
    Format:
    {
      "products": {
        "<product_url>": {
          "status": "partial" | "full_in_stock" | "full_oos",
          "available": [...],
          "unavailable": [...]
        },
        ...
      }
    }
    """
    if not os.path.exists(STATE_FILE):
        return {}

    try:
        with open(STATE_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data.get("products", {})
    except Exception as e:
        print(f"Could not read {STATE_FILE}: {e}")
        return {}


def save_current_state(current_state):
    """
    Save current snapshot to state.json.
    """
    data = {"products": current_state}
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def extract_state_from_report(report):
    """
    Convert report format to simple state:
    {
      product_url: {
        "status": "partial" | "full_in_stock" | "full_oos",
        "available": [...],
        "unavailable": [...]
      }
    }
    """
    state = {}
    for status_key in ("partial", "full_in_stock", "full_oos"):
        bucket = report.get(status_key, {})
        for url, data in bucket.items():
            state[url] = {
                "status": status_key,
                "available": list(data.get("available_sizes", [])),
                "unavailable": list(data.get("unavailable_sizes", [])),
            }
    return state


def has_changes(prev_state, curr_state):
    """
    Return True if there is any change between previous and current states.
    """
    return prev_state != curr_state


# ====== EMAIL SENDING ======

def send_email(subject: str, html_body: str, text_body: str | None = None, attachment_path: str | None = None):
    if not SMTP_USER or not SMTP_PASSWORD:
        raise RuntimeError("SMTP_USER and SMTP_PASSWORD must be set in environment variables.")

    msg = MIMEMultipart("mixed")

    msg["From"] = SMTP_USER

    recipients = [e.strip() for e in TO_EMAIL.split(",") if e.strip()]
    msg["To"] = ", ".join(recipients)

    msg["Subject"] = subject

    # Email body section
    body = MIMEMultipart("alternative")

    if text_body is None:
        text_body = "Please find the attached report."

    body.attach(MIMEText(text_body, "plain"))
    body.attach(MIMEText(html_body, "html"))

    msg.attach(body)

    # Attach Excel file
    if attachment_path and os.path.exists(attachment_path):

        with open(attachment_path, "rb") as attachment:

            part = MIMEBase(
                "application",
                "vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            part.set_payload(attachment.read())

        encoders.encode_base64(part)

        part.add_header(
            "Content-Disposition",
            f'attachment; filename="{os.path.basename(attachment_path)}"'
        )

        msg.attach(part)

    # Send email
    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
        server.starttls()
        server.login(SMTP_USER, SMTP_PASSWORD)
        server.sendmail(SMTP_USER, recipients, msg.as_string())


# ====== EXCEL EXPORT ======

def generate_excel_report(output_filename: str = "CAVA_Stock_Report.xlsx"):
    """
    Generate an Excel spreadsheet with stock data in the desired format.
    Columns:
    - Product Name
    - Product MRP
    - Discount Price
    - Discount Percentage
    - Available Sizes List
    - Zero Sizes
    - Status (Broken/Non Broken)
    - Broken Sizes
    - Non-Broken Sizes Inventory
    """
    try:
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed. Skipping Excel generation.")
        return

    products = fetch_all_products()
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Report"

    # Define headers
    headers = [
        "Product Name",
        "Product MRP",
        "Discount Price",
        "Discount Percentage",
        "Unavailable Sizes",    
        "Available Sizes"
    ]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 25
    
  

    # Define colors
    broken_fill = PatternFill(start_color="FF6961", end_color="FF6961", fill_type="solid")  # Red
    non_broken_fill = PatternFill(start_color="77DD77", end_color="77DD77", fill_type="solid")  # Green
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Process each product
    row_num = 2
    for p in products:
        title = p.get("title", "Unknown product")
        handle = p.get("handle", "")
        
        # Get variants and calculate inventory/pricing
        variants = p.get("variants", [])
        
        available_sizes = []
        unavailable_sizes = []
        total_inventory = 0
        available_inventory = 0
        mrp = None
        discount_price = None
        discount_percentage = None

        # Extract pricing and inventory from first variant
        if variants and len(variants) > 0:
            first_variant = variants[0]
            # MRP is typically stored as compare_at_price
            mrp = first_variant.get("compare_at_price") or first_variant.get("price", "N/A")
            # Discount price is the regular price
            discount_price = first_variant.get("price", "N/A")
            
            # Calculate discount percentage
            if mrp and discount_price:
                try:
                    mrp_float = float(mrp)
                    price_float = float(discount_price)
                    if mrp_float > 0:
                        discount_percentage = round(((mrp_float - price_float) / mrp_float) * 100, 2)
                except (ValueError, TypeError):
                    discount_percentage = "N/A"

        for v in variants:
            size = v.get("option1") or v.get("title", "")
            available = v.get("available", True)
            inventory = v.get("inventory_quantity", 0)
            
            total_inventory += inventory
            
            if available:
                available_sizes.append(size)
                available_inventory += max(inventory, 1)
            else:
                unavailable_sizes.append(size)

        zero_sizes = len(unavailable_sizes)
        
        # Determine status
        if len(available_sizes) == 0:
            # All sizes sold out
            status = "Broken"
            status_fill = broken_fill
        else:
            # At least one size available
            status = "Non Broken"
            status_fill = non_broken_fill

        broken_sizes_str = ", ".join(sorted(set(unavailable_sizes))) if unavailable_sizes else "None"
        non_broken_sizes_str = ", ".join(sorted(set(available_sizes))) if available_sizes else "None"

        # Write row data
        ws.cell(row=row_num, column=1).value = title
        ws.cell(row=row_num, column=2).value = mrp
        ws.cell(row=row_num, column=3).value = discount_price
        ws.cell(row=row_num, column=4).value = discount_percentage
        ws.cell(row=row_num, column=5).value = broken_sizes_str
        ws.cell(row=row_num, column=6).value = non_broken_sizes_str

        # Apply formatting to all cells in row
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        row_num += 1

    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save the workbook
    wb.save(output_filename)
    print(f"Excel report saved: {output_filename}")
    return output_filename


# ====== EMAIL SENDING ======

def main():
    # Generate Excel report (always)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    excel_filename = f"CAVA_Stock_Report_{timestamp}.xlsx"
    generate_excel_report(excel_filename)
    
    # Build today's full report (partial + full in stock + full OOS)
    report = build_report_via_products_json()
    html_body = """
    <html>
    <body style="font-family:Arial, sans-serif;">
        <h2>CAVA Stock Report</h2>
        <p>Please find the attached Excel stock report.</p>
    </body>
    </html>
    """

    text_body = "Please find the attached Excel stock report."

    # Build simple state from report
    curr_state = extract_state_from_report(report)
    prev_state = load_previous_state()

    changed = has_changes(prev_state, curr_state)
    print(f"State changed since last run: {changed}")

    # Save current state for next run (always)
    save_current_state(curr_state)

    # Decide whether to email
    if ONLY_EMAIL_IF_CHANGES and not changed:
        print("No change in availability since last run; not sending email.")
        return

    subject = "CAVA Daily Stock Report (Full Inventory View)"
    send_email(subject, html_body, text_body, excel_filename)
    print("Email sent.")


if __name__ == "__main__":
    main()
